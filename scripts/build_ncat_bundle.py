"""Assemble the NCAT submission bundle: 90-day Excel + 1-year forecast + forms.

Produces, under data/exports/ncat_<rundate>/:
  - Budget_2026-03..06_*.xlsx       monthly budget workbooks (rolling-90d window)
  - FinMg_NCAT_Summary.xlsx         3-month category summary + Section D forecast
  - annual_accounts_<period>.pdf    NSWTG Annual Accounts (past actuals)
  - plan_<period>.pdf               NSWTG Private Manager's Plan (forward forecast)
  - MANIFEST.txt                    human-readable cover sheet (incl. gap/compliance status)
  - NCAT_Submission_<rundate>.zip   everything above, zipped

Read-mostly: writes files to data/exports only; does NOT touch the submissions
register or audit log (do the official register-save + lodgement in the UI so
it is attributed to Linda). Run:

  source .venv/bin/activate && python3 scripts/build_ncat_bundle.py --rundate 2026-06-20
"""

from __future__ import annotations

import argparse
import sys
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from src.db.database import get_connection, init_db
from src.db.queries import get_all_transactions, get_distinct_months, rows_to_transactions
from src.db.queries_estate import bootstrap_managed_person_if_empty
from src.pipeline.excel_writer import write_budget_excel
from src.pipeline.merger import merge_accounts
from src.services.artifacts.fill import fill_artifact
from src.services.artifacts.resolvers import Ctx
from src.services.artifacts.spec import load_spec
from src.services.audit import audit_artifact
from src.services.compliance.engine import evaluate_compliance

REPO = Path(__file__).resolve().parent.parent
WINDOW_MONTHS = ["2026-03", "2026-04", "2026-05", "2026-06"]  # rolling ~90d window
HEADER_FILL = PatternFill("solid", fgColor="EAE3D4")
HEADER_FONT = Font(bold=True, color="1F2A2E")


def _month_workbooks(conn, out_dir: Path) -> list[Path]:
    paths = []
    for month in WINDOW_MONTHS:
        rows = get_all_transactions(conn, month=month)
        if not rows:
            continue
        txns = rows_to_transactions(rows)
        by_account: dict[str, dict[str, list]] = {}
        for t in txns:
            by_account.setdefault(t.account_number, {}).setdefault(month, []).append(t)
        merged = merge_accounts(by_account)
        if month in merged:
            paths.append(Path(write_budget_excel(merged[month], month, str(out_dir))))
    return paths


def _rent_figures(conn, mp_id: int, period_start: str) -> dict:
    """Pull Renato's rent picture from the DB for the cover sheet."""
    actual = conn.execute(
        "SELECT ROUND(COALESCE(SUM(withdrawal),0),2) v FROM transactions "
        "WHERE category='Rent' AND month IN ('2026-03','2026-04','2026-05','2026-06')"
    ).fetchone()["v"]
    arrears = conn.execute(
        "SELECT COALESCE(SUM(amount),0) v FROM debts_liabilities WHERE managed_person_id=?",
        (mp_id,),
    ).fetchone()["v"]
    fc = conn.execute(
        "SELECT f.forecast_value v FROM forecasts f JOIN forecast_categories fc ON fc.id=f.category_id "
        "WHERE f.managed_person_id=? AND f.period_start=? AND fc.category_name='Rent'",
        (mp_id, period_start),
    ).fetchone()
    return {"actual": actual, "arrears": arrears, "forecast": fc["v"] if fc else 0.0}


def _cover_sheet(ws, conn, mp_id: int, period_start: str) -> None:
    """Linda → NCAT cover page, incl. an explicit Rent summary."""
    rent = _rent_figures(conn, mp_id, period_start)
    title = Font(bold=True, size=15, color="1F2A2E")
    h = Font(bold=True, size=12, color="2F6B60")
    b = Font(bold=True)
    rows = [
        ("Annual Financial Report & Forward Plan", title),
        ("Submitted to the NSW Civil and Administrative Tribunal (NCAT)", Font(size=11)),
        ("c/- NSW Trustee & Guardian", Font(size=10, italic=True)),
        ("", None),
        ("Submitted by:", h),
        ("Linda Jane Travia — Private Financial Manager", b),
        ("Lifelong partner of the managed person", Font(italic=True)),
        ("", None),
        ("On behalf of (managed person):", h),
        ("Renato Gentili", b),
        ("136 Madeline St, Belfield NSW 2191", None),
        ("", None),
        ("Reporting period (actuals):", h),
        ("1 March 2026 – 15 June 2026", None),
        ("Forward plan period:", h),
        (f"{period_start} – 12 months", None),
        (f"Prepared: {datetime.now().strftime('%d %B %Y')}", Font(italic=True)),
        ("", None),
        ("RENT — Renato's share", h),
        (f"Rent: $1,000/month, paid from Renato's account to GOODWILL CARE → ${rent['forecast']:,.2f}/yr", b),
        (f"Rent paid within this reporting period: ${rent['actual']:,.2f}", None),
        ("", None),
        ("Declaration:", h),
        ("This report and forward plan have been prepared by Linda Jane Travia in her", None),
        ("capacity as Private Financial Manager for Renato Gentili, from his ANZ bank", None),
        ("statements (3 accounts), for submission to NCAT / NSW Trustee & Guardian.", None),
    ]
    for text, font in rows:
        ws.append([text])
        if font:
            ws[ws.max_row][0].font = font
    ws.column_dimensions["A"].width = 95


def _summary_workbook(conn, mp_id: int, period_start: str, out_dir: Path) -> Path:
    wb = Workbook()

    # Sheet 1 — Linda → NCAT cover page.
    _cover_sheet(wb.active, conn, mp_id, period_start)
    wb.active.title = "Cover"

    # Sheet 2 — 3-month category actuals (the rolling window).
    ws = wb.create_sheet("3-Month Actuals")
    ws.append(["FinMg — Rolling 90-Day Category Summary (Mar–Jun 2026)"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(["Category", "Type", "Withdrawals", "Deposits", "Txns"])
    for c in ws[3]:
        c.fill, c.font = HEADER_FILL, HEADER_FONT
    rows = conn.execute(
        """
        SELECT category,
               ROUND(SUM(COALESCE(withdrawal,0)),2) w,
               ROUND(SUM(COALESCE(deposit,0)),2) d,
               COUNT(*) n
        FROM transactions
        WHERE month IN ('2026-03','2026-04','2026-05','2026-06')
          AND is_internal_transfer = 0
        GROUP BY category ORDER BY w DESC, d DESC
        """
    ).fetchall()
    tot_w = tot_d = 0.0
    for r in rows:
        ws.append([r["category"], "expense" if r["w"] else "income", r["w"], r["d"], r["n"]])
        tot_w += r["w"] or 0
        tot_d += r["d"] or 0
    ws.append([])
    ws.append(["TOTAL", "", round(tot_w, 2), round(tot_d, 2), ""])
    ws[ws.max_row][0].font = Font(bold=True)
    for col, width in zip("ABCDE", (30, 12, 14, 14, 8)):
        ws.column_dimensions[col].width = width

    # Sheet 2 — Section D 1-year forecast.
    ws2 = wb.create_sheet("1-Year Forecast (Sec D)")
    ws2.append(["FinMg — Annualized Section D Forecast (next 12 months)"])
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.append([])
    ws2.append(["Section", "Category", "Trailing actual", "Forecast (1yr)", "Basis"])
    for c in ws2[3]:
        c.fill, c.font = HEADER_FILL, HEADER_FONT
    frows = conn.execute(
        """
        SELECT fc.section, fc.category_name, f.actual_value, f.forecast_value, f.override_reason
        FROM forecasts f JOIN forecast_categories fc ON fc.id = f.category_id
        WHERE f.managed_person_id=? AND f.period_start=?
        ORDER BY fc.section, fc.display_order
        """,
        (mp_id, period_start),
    ).fetchall()
    inc = exp = 0.0
    for r in frows:
        basis = (r["override_reason"] or "").replace(" [auto]", "")
        ws2.append([
            r["section"], r["category_name"],
            None if r["actual_value"] is None else round(r["actual_value"], 2),
            None if r["forecast_value"] is None else round(r["forecast_value"], 2),
            basis,
        ])
        if r["forecast_value"]:
            if r["section"] == "D_income":
                inc += r["forecast_value"]
            else:
                exp += r["forecast_value"]
    ws2.append([])
    ws2.append(["", "FORECAST INCOME (1yr)", "", round(inc, 2), ""])
    ws2.append(["", "FORECAST EXPENDITURE (1yr)", "", round(exp, 2), ""])
    ws2.append(["", "NET (income − expenditure)", "", round(inc - exp, 2), ""])
    for i in range(2, -1, -1):
        ws2[ws2.max_row - i][1].font = Font(bold=True)
    for col, width in zip("ABCDE", (14, 28, 16, 16, 70)):
        ws2.column_dimensions[col].width = width

    path = out_dir / "FinMg_NCAT_Summary.xlsx"
    wb.save(path)
    return path


def _artifact_pdf(conn, mp_id, key, p_start, p_end, out_dir):
    spec = load_spec(key)
    ctx = Ctx(conn=conn, managed_person_id=mp_id, period_start=p_start, period_end=p_end)
    report = audit_artifact(conn, spec, ctx)
    compliance = evaluate_compliance(conn, mp_id, p_start, p_end)
    filled = fill_artifact(spec, ctx)
    path = out_dir / f"{key}_{p_start}_{p_end}.pdf"
    path.write_bytes(filled.pdf_bytes)
    return path, report, compliance, filled


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--rundate", default=date.today().isoformat())
    args = ap.parse_args()
    rundate = args.rundate

    out_dir = REPO / "data" / "exports" / f"ncat_{rundate.replace('-', '')}"
    out_dir.mkdir(parents=True, exist_ok=True)

    conn = get_connection()
    init_db(conn)
    mp_id = bootstrap_managed_person_if_empty(conn, "GENTILI", "Renato")

    today = date.fromisoformat(rundate)
    plan_start, plan_end = today.isoformat(), (today.replace(year=today.year + 1) - timedelta(days=1)).isoformat()
    aa_start, aa_end = "2026-03-01", "2026-06-15"  # actuals window (rolling ~90d, full statement coverage)

    print("→ monthly budget workbooks…")
    month_paths = _month_workbooks(conn, out_dir)
    print("→ summary workbook…")
    summary_path = _summary_workbook(conn, mp_id, plan_start, out_dir)
    print("→ annual accounts PDF…")
    aa_path, aa_rep, aa_cmp, aa_fill = _artifact_pdf(conn, mp_id, "annual_accounts", aa_start, aa_end, out_dir)
    print("→ plan PDF…")
    pl_path, pl_rep, pl_cmp, pl_fill = _artifact_pdf(conn, mp_id, "plan", plan_start, plan_end, out_dir)
    conn.close()

    # MANIFEST
    lines = [
        "NCAT SUBMISSION — Annual Financial Report & Forward Plan",
        "Submitted to: NSW Civil and Administrative Tribunal (NCAT) c/- NSW Trustee & Guardian",
        "Submitted by: Linda Jane Travia — Private Financial Manager",
        "On behalf of: Renato Gentili (managed person), 136 Madeline St, Belfield NSW 2191",
        f"Prepared: {datetime.now().strftime('%d %B %Y, %H:%M')}",
        "",
        "CONTENTS",
        "--------",
        "Monthly budget workbooks (rolling 90-day window, Mar–Jun 2026):",
    ]
    lines += [f"  - {p.name}" for p in month_paths]
    lines += [
        f"  - {summary_path.name}   (3-month category summary + 1-year forecast)",
        "",
        f"NSWTG Annual Accounts (actuals {aa_start}..{aa_end}):",
        f"  - {aa_path.name}",
        f"      completeness {aa_rep.completeness*100:.0f}% · {len(aa_rep.gaps)} gap(s) · "
        f"{len(aa_fill.resolved)} fields filled · {len(aa_cmp.blocking)} compliance block(s)",
        "",
        f"NSWTG Private Manager's Plan (forecast {plan_start}..{plan_end}):",
        f"  - {pl_path.name}",
        f"      completeness {pl_rep.completeness*100:.0f}% · {len(pl_rep.gaps)} gap(s) · "
        f"{len(pl_fill.resolved)} fields filled · {len(pl_cmp.blocking)} compliance block(s)",
        "",
        "OPEN GAPS (fill in Identity / Inventory / Forecast, or record N/A rationale):",
    ]
    for label, rep in (("Annual Accounts", aa_rep), ("Plan", pl_rep)):
        if rep.gaps:
            lines.append(f"  {label}: {', '.join(rep.gaps)}")
        else:
            lines.append(f"  {label}: none")
    for label, cmp in (("Annual Accounts", aa_cmp), ("Plan", pl_cmp)):
        if cmp.blocking:
            lines.append(f"  COMPLIANCE BLOCKS — {label}:")
            lines += [f"    - {g.finding.handbook_ref} {g.finding.title}" for g in cmp.blocking]
    manifest = out_dir / "MANIFEST.txt"
    manifest.write_text("\n".join(lines) + "\n")

    # ZIP
    zip_path = out_dir / f"NCAT_Submission_{rundate.replace('-', '')}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for p in [*month_paths, summary_path, aa_path, pl_path, manifest]:
            z.write(p, p.name)

    print("\n" + "\n".join(lines))
    print(f"\n✅ Bundle: {zip_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
