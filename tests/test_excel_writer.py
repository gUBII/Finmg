"""Tests for Excel budget export."""

import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent))

from src.models.transaction import Transaction
from src.pipeline.excel_writer import write_budget_excel


class TestExcelWriter:

    def test_excludes_internal_transfer_category_and_clears_template_samples(self, tmp_path):
        transactions = [
            Transaction(
                date=date(2026, 1, 2),
                description="WOOLWORTHS/174 LAKEMBA ST LAKEMBA",
                withdrawal=87.41,
                category="Groceries",
                account_number="178865319",
            ),
            Transaction(
                date=date(2026, 1, 29),
                description="ANZ M-BANKING FUNDS TFER TRANSFER 158182 FROM 178873749",
                deposit=800.0,
                category="Internal Transfer",
                is_internal_transfer=False,
                account_number="178865319",
            ),
        ]

        filepath = write_budget_excel(transactions, "2026-01", str(tmp_path))

        wb = load_workbook(filepath)
        ws = wb["Transactions"]

        assert ws["B5"].value.date().isoformat() == "2026-01-02"
        assert ws["C5"].value == 87.41
        assert ws["D5"].value == "WOOLWORTHS/174 LAKEMBA ST LAKEMBA"
        assert ws["E5"].value == "Groceries"

        # The template's sample income row should be cleared when there is no real income.
        assert ws["G5"].value is None
        assert ws["H5"].value is None
        assert ws["I5"].value is None
        assert ws["J5"].value is None
