"""Reproducible Excel export of the §76 gift-forecast matrix.

Renders the recipients × occasions table (the digital twin of the handwritten
gift forecast) styled to the FinMg theme: teal header, serif title, flagged
rows tinted, column-total + grand-total footer. Reads the ledger via
src/services/gift_forecast.build_matrix so the workbook always matches the app.
"""
from __future__ import annotations

import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.services.gift_forecast import OCC_LABEL, OCC_ORDER, TITLE, build_matrix

# FinMg theme tokens (calm-ledger teal/serif)
TEAL = "1F4E5F"
HEADER_TEXT = "FFFFFF"
FLAG_FILL = "FCE8E6"
TOTAL_FILL = "EDF3F4"
GRID = "C9D6D9"
SERIF = "Georgia"


def write_gift_forecast_xlsx(
    conn: sqlite3.Connection, managed_person_id: int, path: str | Path, title: str = TITLE
) -> tuple[Path, float]:
    """Write the styled gift-forecast matrix to `path`. Returns (path, grand_total)."""
    rows, col_totals, grand = build_matrix(conn, managed_person_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Gift Forecast"
    thin = Side(style="thin", color=GRID)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor=TEAL)
    flag_fill = PatternFill("solid", fgColor=FLAG_FILL)
    total_fill = PatternFill("solid", fgColor=TOTAL_FILL)
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ncols = 2 + len(OCC_ORDER) + 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(name=SERIF, bold=True, size=13, color=TEAL)
    t.alignment = Alignment(horizontal="left", vertical="center")

    headers = ["Person's Name", "Relation"] + [OCC_LABEL[o] for o in OCC_ORDER] + ["Row Total"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = Font(name=SERIF, bold=True, color=HEADER_TEXT)
        cell.fill = header_fill
        cell.alignment = centre
        cell.border = border

    r = 4
    for row in rows:
        ws.cell(row=r, column=1, value=row.name).border = border
        ws.cell(row=r, column=2, value=row.relation).border = border
        for i, occ in enumerate(OCC_ORDER, 3):
            amt = row.amounts.get(occ)
            cell = ws.cell(row=r, column=i, value=(round(amt, 2) if amt else None))
            cell.alignment = centre
            cell.border = border
            cell.number_format = "$#,##0"
        rt = ws.cell(row=r, column=ncols, value=round(row.total, 2))
        rt.font = Font(bold=True)
        rt.alignment = centre
        rt.border = border
        rt.number_format = "$#,##0"
        if row.flagged:
            for c in range(1, ncols + 1):
                ws.cell(row=r, column=c).fill = flag_fill
        for c in range(1, 3):
            ws.cell(row=r, column=c).font = Font(name=SERIF, bold=(c == 1))
        r += 1

    foot = ws.cell(row=r, column=1, value="Column total")
    foot.font = Font(name=SERIF, bold=True)
    for i, occ in enumerate(OCC_ORDER, 3):
        cell = ws.cell(row=r, column=i, value=round(col_totals[occ], 2))
        cell.font = Font(bold=True)
        cell.alignment = centre
        cell.number_format = "$#,##0"
    g = ws.cell(row=r, column=ncols, value=round(grand, 2))
    g.font = Font(name=SERIF, bold=True, color=TEAL)
    g.alignment = centre
    g.number_format = "$#,##0"
    for c in range(1, ncols + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = total_fill
        cell.border = border

    widths = [20, 18] + [12] * len(OCC_ORDER) + [12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=3, column=i).column_letter].width = w
    ws.row_dimensions[3].height = 30
    ws.freeze_panes = "C4"

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path, grand
