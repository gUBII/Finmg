"""
Generate budget Excel workbooks by populating Renato's monthly template.

The template has:
- Summary sheet: formulas auto-calculate from Transactions sheet
- Transactions sheet: Expenses (cols B-E) and Income (cols G-J)
  Row 4: Headers (Date, Amount, Description, Category)
  Row 5+: Data rows

This module copies the template and populates the Transactions sheet.
The Summary sheet formulas (SUMIF) auto-update based on the data.
"""

import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from src.models.transaction import Transaction

TEMPLATE_PATH = Path(__file__).parent.parent.parent / "template_empty" / "Monthly Template - Renato Gentili.xlsx"

MONTH_NAMES = {
    "01": "January", "02": "February", "03": "March", "04": "April",
    "05": "May", "06": "June", "07": "July", "08": "August",
    "09": "September", "10": "October", "11": "November", "12": "December",
}

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def _is_excluded_transfer(txn: Transaction) -> bool:
    """Treat explicit transfer categories as excluded even if the flag is stale."""
    return txn.is_internal_transfer or txn.category == "Internal Transfer"


def write_budget_excel(
    transactions: list[Transaction],
    month_key: str,
    output_dir: str,
) -> str:
    """
    Write a budget Excel workbook for one month by populating the template.

    Internal transfers are excluded from the budget.
    Withdrawals go into the Expenses section (cols B-E).
    Deposits go into the Income section (cols G-J).
    """
    month_name = MONTH_NAMES.get(month_key.split("-")[1], month_key)
    filename = f"Budget_{month_key}_{month_name}.xlsx"
    filepath = Path(output_dir) / filename
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    # Copy template
    if TEMPLATE_PATH.exists():
        shutil.copy2(str(TEMPLATE_PATH), str(filepath))
        wb = load_workbook(str(filepath))
    else:
        # Fallback: create from scratch if template not found
        return _write_budget_excel_fallback(transactions, month_key, output_dir)

    ws = wb["Transactions"]

    # Clear the template sample rows before writing actual data.
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=2, max_col=10):
        for cell in row:
            cell.value = None
            cell.fill = PatternFill(fill_type=None)

    # Separate expenses and income (exclude internal transfers)
    expenses = [t for t in transactions if t.withdrawal and not _is_excluded_transfer(t)]
    income = [t for t in transactions if t.deposit and not _is_excluded_transfer(t)]

    # Sort by date
    expenses.sort(key=lambda t: t.date)
    income.sort(key=lambda t: t.date)

    # Write expenses starting at row 5 (cols B=2, C=3, D=4, E=5)
    for i, txn in enumerate(expenses):
        row = 5 + i
        ws.cell(row=row, column=2, value=datetime(txn.date.year, txn.date.month, txn.date.day))
        ws.cell(row=row, column=2).number_format = "DD/MM/YYYY"
        ws.cell(row=row, column=3, value=txn.withdrawal)
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=_clean_description(txn.description))
        ws.cell(row=row, column=5, value=txn.category)

        # Highlight uncategorised
        if txn.category == "Uncategorised":
            for c in range(2, 6):
                ws.cell(row=row, column=c).fill = YELLOW_FILL

    # Write income starting at row 5 (cols G=7, H=8, I=9, J=10)
    for i, txn in enumerate(income):
        row = 5 + i
        ws.cell(row=row, column=7, value=datetime(txn.date.year, txn.date.month, txn.date.day))
        ws.cell(row=row, column=7).number_format = "DD/MM/YYYY"
        ws.cell(row=row, column=8, value=txn.deposit)
        ws.cell(row=row, column=8).number_format = '#,##0.00'
        ws.cell(row=row, column=9, value=_clean_description(txn.description))
        ws.cell(row=row, column=10, value=txn.category)

    wb.save(str(filepath))
    return str(filepath)


def _clean_description(desc: str) -> str:
    """Clean up transaction description for display."""
    # Remove common prefixes
    for prefix in ["EFTPOS ", "VISA DEBIT PURCHASE CARD 1216 ", "VISA DEBIT PURCHASE CARD 6621 "]:
        if desc.upper().startswith(prefix):
            desc = desc[len(prefix):]
    return desc.strip()


def _write_budget_excel_fallback(
    transactions: list[Transaction],
    month_key: str,
    output_dir: str,
) -> str:
    """Fallback: create Excel from scratch if template not available."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter

    month_name = MONTH_NAMES.get(month_key.split("-")[1], month_key)
    filename = f"Budget_{month_key}_{month_name}.xlsx"
    filepath = Path(output_dir) / filename
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BOLD_FONT = Font(bold=True, size=11)
    CURRENCY_FORMAT = '$#,##0.00'
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    wb = Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Budget Summary — {month_name} {month_key.split('-')[0]}"
    ws["A1"].font = Font(bold=True, size=14)

    # Expenses
    ws["A3"] = "EXPENSES"
    ws["A3"].font = BOLD_FONT
    for col, h in enumerate(["Category", "Count", "Total"], 2):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    categories: dict[str, dict] = {}
    for txn in transactions:
        cat = txn.category
        if cat not in categories:
            categories[cat] = {"count": 0, "withdrawals": 0.0, "deposits": 0.0}
        categories[cat]["count"] += 1
        if not _is_excluded_transfer(txn):
            categories[cat]["withdrawals"] += txn.withdrawal or 0
            categories[cat]["deposits"] += txn.deposit or 0

    row = 5
    total_exp = 0.0
    for cat in sorted(categories):
        data = categories[cat]
        if data["withdrawals"] > 0:
            ws.cell(row=row, column=2, value=cat)
            ws.cell(row=row, column=3, value=data["count"])
            cell = ws.cell(row=row, column=4, value=data["withdrawals"])
            cell.number_format = CURRENCY_FORMAT
            total_exp += data["withdrawals"]
            if cat == "Uncategorised":
                for c in range(2, 5):
                    ws.cell(row=row, column=c).fill = YELLOW_FILL
            row += 1

    row += 1
    ws.cell(row=row, column=2, value="TOTAL EXPENSES").font = BOLD_FONT
    ws.cell(row=row, column=4, value=round(total_exp, 2)).number_format = CURRENCY_FORMAT

    # Income
    row += 2
    ws.cell(row=row, column=2, value="INCOME").font = BOLD_FONT
    row += 1
    for col, h in enumerate(["Category", "Count", "Total"], 2):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")

    row += 1
    total_inc = 0.0
    for cat in sorted(categories):
        data = categories[cat]
        if data["deposits"] > 0:
            ws.cell(row=row, column=2, value=cat)
            ws.cell(row=row, column=3, value=data["count"])
            ws.cell(row=row, column=4, value=data["deposits"]).number_format = CURRENCY_FORMAT
            total_inc += data["deposits"]
            row += 1

    row += 1
    ws.cell(row=row, column=2, value="TOTAL INCOME").font = BOLD_FONT
    ws.cell(row=row, column=4, value=round(total_inc, 2)).number_format = CURRENCY_FORMAT

    row += 2
    ws.cell(row=row, column=2, value="NET POSITION").font = Font(bold=True, size=12)
    net = round(total_inc - total_exp, 2)
    cell = ws.cell(row=row, column=4, value=net)
    cell.number_format = CURRENCY_FORMAT
    cell.font = Font(bold=True, size=12, color="00B050" if net >= 0 else "FF0000")

    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 15

    # Transactions sheet
    ws2 = wb.create_sheet("Transactions")
    headers = ["Date", "Description", "Account", "Account Type", "Category", "Withdrawal", "Deposit", "Internal Transfer"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for row_idx, txn in enumerate(transactions, 2):
        ws2.cell(row=row_idx, column=1, value=txn.date.isoformat())
        ws2.cell(row=row_idx, column=2, value=txn.description)
        ws2.cell(row=row_idx, column=3, value=txn.account_number)
        ws2.cell(row=row_idx, column=4, value=txn.account_type)
        ws2.cell(row=row_idx, column=5, value=txn.category)
        if txn.withdrawal:
            ws2.cell(row=row_idx, column=6, value=txn.withdrawal).number_format = CURRENCY_FORMAT
        if txn.deposit:
            ws2.cell(row=row_idx, column=7, value=txn.deposit).number_format = CURRENCY_FORMAT
        ws2.cell(row=row_idx, column=8, value="Yes" if txn.is_internal_transfer else "")
        if txn.category == "Uncategorised":
            for c in range(1, 9):
                ws2.cell(row=row_idx, column=c).fill = YELLOW_FILL

    widths = [12, 50, 12, 22, 25, 14, 14, 16]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(str(filepath))
    return str(filepath)
